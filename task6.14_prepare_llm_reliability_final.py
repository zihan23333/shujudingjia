#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import re
import shutil
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
PKG = ROOT / "annotation_package_final"
REP_DIR = PKG / "representative"
HARD_DIR = PKG / "hardcase"
INPUT_CONTEXTS = ROOT / "target_aligned_contexts_final.csv"
INPUT_LLM = ROOT / "llm_results_target_aligned_final.csv"

CUE_WORDS = [
    "however",
    "limitation",
    "limited",
    "fail",
    "fails",
    "failed",
    "unlike",
    "drawback",
    "problem",
    "cannot",
    "insufficient",
    "challenge",
    "although",
    "but",
    "weakness",
    "shortcoming",
]


def normalize_section(value: object) -> str:
    text = str(value).strip().lower()
    if "method" in text:
        return "Methodology"
    if "result" in text:
        return "Result"
    if "discussion" in text:
        return "Discussion"
    if "conclusion" in text:
        return "Conclusion"
    if "introduction" in text or "related" in text or "background" in text:
        return "Introduction"
    return "Other"


def slugify(text: object, max_len: int = 48) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(text or "").strip())
    s = re.sub(r"_+", "_", s).strip("_")
    return (s[:max_len] or "untitled")


def control_clean(value: object) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", str(value))


def all_pdf_files() -> List[Path]:
    pdfs = []
    for p in ROOT.rglob("*.pdf"):
        if "annotation_package_final" in str(p):
            continue
        pdfs.append(p)
    return pdfs


def all_text_files() -> List[Path]:
    txts = []
    for p in ROOT.rglob("*.txt"):
        if "annotation_package_final" in str(p):
            continue
        if p.name.lower().startswith("readme"):
            continue
        txts.append(p)
    return txts


def build_file_indices(files: Iterable[Path]) -> Tuple[Dict[str, Path], Dict[str, List[Path]]]:
    by_id: Dict[str, Path] = {}
    by_norm_title: Dict[str, List[Path]] = {}
    for p in files:
        stem = p.stem
        m = re.search(r"(W\d{6,})", stem)
        if m and m.group(1) not in by_id:
            by_id[m.group(1)] = p
        norm = re.sub(r"[^a-z0-9]+", "", stem.lower())
        if norm:
            by_norm_title.setdefault(norm, []).append(p)
    return by_id, by_norm_title


def find_file_for_paper(paper_id: str, title: str, id_map: Dict[str, Path], title_map: Dict[str, List[Path]]) -> Optional[Path]:
    if paper_id in id_map:
        return id_map[paper_id]
    norm_title = re.sub(r"[^a-z0-9]+", "", str(title).lower())
    if norm_title in title_map:
        return title_map[norm_title][0]
    for norm, paths in title_map.items():
        if norm_title and (norm_title in norm or norm in norm_title):
            return paths[0]
    return None


def add_buckets(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    work["section_bucket"] = work["LLM_section"].apply(normalize_section)
    work["relevance_bucket"] = pd.cut(work["LLM_relevance"], bins=[-0.01, 0.4, 0.7, 1.01], labels=["low", "mid", "high"])
    work["q_bucket"] = pd.cut(work["LLM_q"], bins=[-0.01, 0.08, 0.22, 1.01], labels=["low", "mid", "high"])
    work["stratum"] = (
        work["alignment_status"].astype(str) + "|" +
        work["section_bucket"].astype(str) + "|" +
        work["relevance_bucket"].astype(str) + "|" +
        work["q_bucket"].astype(str)
    )
    return work


def pick_with_source_diversity(df: pd.DataFrame, n: int) -> pd.DataFrame:
    selected: List[pd.Series] = []
    used_edges = set()
    used_sources = set()
    for _, group in df.groupby("stratum", sort=False):
        candidate = group[~group["source_id"].isin(used_sources)]
        row = (candidate if not candidate.empty else group).sort_values(["source_id", "edge_id"]).iloc[0]
        selected.append(row)
        used_edges.add(row["edge_id"])
        used_sources.add(row["source_id"])
    if len(selected) < min(n, len(df)):
        remain = df[~df["edge_id"].isin(used_edges)].copy()
        remain = remain.sort_values(["source_id", "LLM_relevance", "LLM_q"], ascending=[True, False, False])
        for _, row in remain.iterrows():
            if len(selected) >= min(n, len(df)):
                break
            selected.append(row)
            used_edges.add(row["edge_id"])
            used_sources.add(row["source_id"])
    out = pd.DataFrame(selected).drop_duplicates("edge_id").head(min(n, len(df))).reset_index(drop=True)
    return out


def build_representative_sample(df: pd.DataFrame, n: int = 50) -> pd.DataFrame:
    work = add_buckets(df)
    return pick_with_source_diversity(work, n)


def build_hardcase_sample(df: pd.DataFrame, n: int = 30) -> pd.DataFrame:
    pattern = re.compile(r"\b(" + "|".join(re.escape(w) for w in CUE_WORDS) + r")\b", flags=re.I)
    work = df.copy()
    work["matched_cue_words"] = work["all_target_aligned_contexts"].fillna("").apply(
        lambda s: ", ".join(sorted(set(m.group(0).lower() for m in pattern.finditer(str(s)))))
    )
    work = work[work["matched_cue_words"].astype(str).str.strip().ne("")].copy()
    if work.empty:
        return work
    work["cue_count"] = work["matched_cue_words"].apply(lambda s: len([x for x in str(s).split(",") if x.strip()]))
    work = add_buckets(work)
    work = work.sort_values(
        ["cue_count", "alignment_status", "LLM_relevance", "LLM_q"],
        ascending=[False, True, False, False],
    )
    selected = pick_with_source_diversity(work, n)
    return selected


def write_xlsx(df: pd.DataFrame, path: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([control_clean(v) for v in row.values])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    input_cols = {
        "human_alignment_check",
        "human_primary_section",
        "human_sentiment",
        "human_relevance",
        "human_note",
    }
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, name in enumerate(df.columns, start=1):
        width = 18
        if name in {"citing_paper_title", "cited_paper_title", "target_reference_entry"}:
            width = 34
        if name in {"all_target_aligned_contexts", "human_note"}:
            width = 60
        if name.endswith("_path"):
            width = 44
        ws.column_dimensions[get_column_letter(idx)].width = width
        if name in input_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, idx).fill = input_fill

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 52
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(path)


def enrich_and_copy(
    df: pd.DataFrame,
    package_dir: Path,
    pdf_id_map: Dict[str, Path],
    pdf_title_map: Dict[str, List[Path]],
    text_id_map: Dict[str, Path],
    text_title_map: Dict[str, List[Path]],
) -> Tuple[pd.DataFrame, List[str], List[str]]:
    source_pdf_dir = package_dir / "source_pdfs_by_sample_id"
    source_text_dir = package_dir / "source_texts_by_sample_id"
    target_pdf_dir = package_dir / "target_pdfs_by_sample_id"
    snippet_dir = package_dir / "context_snippets_by_sample_id"
    for d in [source_pdf_dir, source_text_dir, target_pdf_dir, snippet_dir]:
        d.mkdir(parents=True, exist_ok=True)

    source_missing: List[str] = []
    target_missing: List[str] = []
    rows = []
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        sample_id = f"sample{i:03d}"
        source_pdf = find_file_for_paper(str(row["source_id"]), str(row["citing_paper_title"]), pdf_id_map, pdf_title_map)
        target_pdf = find_file_for_paper(str(row["target_id"]), str(row["cited_paper_title"]), pdf_id_map, pdf_title_map)
        source_text = find_file_for_paper(str(row["source_id"]), str(row["citing_paper_title"]), text_id_map, text_title_map)

        source_pdf_filename = ""
        source_pdf_path = ""
        target_pdf_filename = ""
        target_pdf_path = ""
        source_text_path = ""
        source_pdf_missing = 1
        target_pdf_missing = 1

        if source_pdf and source_pdf.exists():
            source_pdf_filename = f"{sample_id}__source__{slugify(row['citing_paper_title'])}.pdf"
            copied = source_pdf_dir / source_pdf_filename
            shutil.copy2(source_pdf, copied)
            source_pdf_path = str(copied.resolve())
            source_pdf_missing = 0
        else:
            source_missing.append(f"{sample_id}: {row['citing_paper_title']}")

        if target_pdf and target_pdf.exists():
            target_pdf_filename = f"{sample_id}__target__{slugify(row['cited_paper_title'])}.pdf"
            copied = target_pdf_dir / target_pdf_filename
            shutil.copy2(target_pdf, copied)
            target_pdf_path = str(copied.resolve())
            target_pdf_missing = 0
        else:
            target_missing.append(f"{sample_id}: {row['cited_paper_title']}")

        if source_text and source_text.exists():
            copied = source_text_dir / f"{sample_id}__source__{slugify(row['citing_paper_title'])}.txt"
            shutil.copy2(source_text, copied)
            source_text_path = str(copied.resolve())

        snippet = snippet_dir / f"{sample_id}_context.md"
        snippet.write_text(
            "\n".join(
                [
                    f"# {sample_id}",
                    "",
                    f"- citing paper title: {row['citing_paper_title']}",
                    f"- cited paper title: {row['cited_paper_title']}",
                    f"- target reference marker: {row['target_reference_marker']}",
                    f"- target reference entry: {row['target_reference_entry']}",
                    f"- alignment status: {row['alignment_status']}",
                    "",
                    "## all_target_aligned_contexts",
                    "",
                    str(row["all_target_aligned_contexts"]),
                    "",
                    "## LLM outputs",
                    "",
                    f"- LLM_section: {row['LLM_section']}",
                    f"- LLM_sentiment: {row['LLM_sentiment']}",
                    f"- LLM_relevance: {row['LLM_relevance']}",
                    f"- LLM_confidence: {row['LLM_confidence']}",
                    f"- LLM_q: {row['LLM_q']}",
                    "",
                    "## Human scoring reminder",
                    "",
                    "- First judge whether the context truly aligns to the current target cited paper.",
                    "- Only score the current target paper even for grouped/range citation contexts.",
                    "- Use `wrong_or_ambiguous` if the context still cannot be confidently aligned to the current target.",
                ]
            ),
            encoding="utf-8",
        )

        row_out = row.copy()
        row_out["sample_id"] = sample_id
        row_out["source_pdf_filename"] = source_pdf_filename
        row_out["source_pdf_path"] = source_pdf_path
        row_out["source_text_path"] = source_text_path
        row_out["target_pdf_filename"] = target_pdf_filename
        row_out["target_pdf_path"] = target_pdf_path
        row_out["source_pdf_missing"] = source_pdf_missing
        row_out["target_pdf_missing"] = target_pdf_missing
        rows.append(row_out)

    enriched = pd.DataFrame(rows)
    front_cols = [
        "sample_id",
        "edge_id",
        "source_id",
        "target_id",
        "citing_paper_title",
        "cited_paper_title",
        "alignment_status",
        "target_reference_marker",
        "target_reference_entry",
        "all_target_aligned_contexts",
        "source_pdf_filename",
        "source_pdf_path",
        "source_text_path",
        "target_pdf_filename",
        "target_pdf_path",
        "source_pdf_missing",
        "target_pdf_missing",
        "LLM_section",
        "LLM_sentiment",
        "LLM_relevance",
        "LLM_confidence",
        "LLM_q",
        "human_alignment_check",
        "human_primary_section",
        "human_sentiment",
        "human_relevance",
        "human_note",
    ]
    keep_cols = [c for c in front_cols if c in enriched.columns]
    return enriched[keep_cols], source_missing, target_missing


def guidelines_text() -> str:
    return """# 人工评分说明（final）

本次人工评分面向正式 final 语义层，所有样本均来自 113 条 DeepSeek target-aligned citation edges。

## 1. human_alignment_check

先判断当前 `all_target_aligned_contexts` 是否确实对应当前 `cited_paper_title`：

- `correct`：上下文明确对应当前 cited paper；
- `wrong_or_ambiguous`：上下文无法确认对应 target，或明显对应其他文献。

只有 `human_alignment_check = correct` 的样本进入正式 reliability 计算。

## 2. human_primary_section

从以下类别中选择：

- `Introduction`
- `Methodology`
- `Result`
- `Discussion`
- `Conclusion`
- `Other`

若 grouped / range citation 中出现多个文献，只评价当前 target cited paper 在该 context 中的语义作用。

## 3. human_sentiment

取值：

- `positive`
- `neutral`
- `negative`

规则：

- `positive`：采用、扩展、支持、作为方法/数据/理论依据；
- `neutral`：背景介绍、一般相关工作、客观陈述；
- `negative`：批评、指出不足、反驳、作为 limitation 或 problem 的例子。

## 4. human_relevance

五档评分：

- `0`：无关或错误匹配
- `0.25`：弱背景性提及
- `0.5`：主题相关但非核心依赖
- `0.75`：明确用于方法、比较、论证或实验设计
- `1.0`：核心方法、数据、模型、结论或理论依赖

## 5. human_note

记录：

- 为什么判断为 `correct` / `wrong_or_ambiguous`
- 是否为 grouped citation
- 是否有转折或负面线索
- 是否需要查看 PDF 才能判断
"""


def build_readme(rep_source_missing: List[str], rep_target_missing: List[str], hard_source_missing: List[str], hard_target_missing: List[str]) -> str:
    lines = [
        "annotation_package_final",
        "",
        "包含两套人工评分材料：",
        "- representative：正式 reliability 用",
        "- hardcase：附加 hard-case 审计用",
        "",
        "使用说明：",
        "- 先看 annotation_guidelines_final.md",
        "- 只在 human_alignment_check = correct 的样本上做正式 reliability",
        "- hard-case 样本不替代 representative 样本",
        "",
        "Representative 缺失 source PDF：",
    ]
    lines.extend(rep_source_missing or ["- 无"])
    lines.extend(["", "Representative 缺失 target PDF："])
    lines.extend(rep_target_missing or ["- 无"])
    lines.extend(["", "Hardcase 缺失 source PDF："])
    lines.extend(hard_source_missing or ["- 无"])
    lines.extend(["", "Hardcase 缺失 target PDF："])
    lines.extend(hard_target_missing or ["- 无"])
    return "\n".join(lines)


def main() -> None:
    for d in [REP_DIR, HARD_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    contexts = pd.read_csv(INPUT_CONTEXTS, encoding="utf-8-sig")
    llm = pd.read_csv(INPUT_LLM, encoding="utf-8-sig")
    merged = contexts.merge(
        llm[
            ["edge_id", "source_id", "target_id", "section", "sentiment", "relevance", "confidence", "q_ij"]
        ].rename(
            columns={
                "section": "LLM_section",
                "sentiment": "LLM_sentiment",
                "relevance": "LLM_relevance",
                "confidence": "LLM_confidence",
                "q_ij": "LLM_q",
            }
        ),
        on=["edge_id", "source_id", "target_id"],
        how="inner",
    )
    merged = merged[merged["alignment_status"].isin(["high_confidence", "grouped", "range"])].copy()

    rep = build_representative_sample(merged, n=50).copy()
    hard = build_hardcase_sample(merged, n=30).copy()

    for df in [rep, hard]:
        df["human_alignment_check"] = ""
        df["human_primary_section"] = ""
        df["human_sentiment"] = ""
        df["human_relevance"] = ""
        df["human_note"] = ""

    pdfs = all_pdf_files()
    txts = all_text_files()
    pdf_id_map, pdf_title_map = build_file_indices(pdfs)
    txt_id_map, txt_title_map = build_file_indices(txts)

    rep_enriched, rep_source_missing, rep_target_missing = enrich_and_copy(rep, REP_DIR, pdf_id_map, pdf_title_map, txt_id_map, txt_title_map)
    hard_enriched, hard_source_missing, hard_target_missing = enrich_and_copy(hard, HARD_DIR, pdf_id_map, pdf_title_map, txt_id_map, txt_title_map)

    rep_csv = ROOT / "sampled_contexts_for_human_annotation_representative_final.csv"
    rep_xlsx = ROOT / "sampled_contexts_for_human_annotation_representative_final.xlsx"
    hard_csv = ROOT / "sampled_contexts_for_human_annotation_hardcase_final.csv"
    hard_xlsx = ROOT / "sampled_contexts_for_human_annotation_hardcase_final.xlsx"

    rep_enriched.to_csv(rep_csv, index=False, encoding="utf-8-sig")
    hard_enriched.to_csv(hard_csv, index=False, encoding="utf-8-sig")
    rep_enriched.to_csv(REP_DIR / rep_csv.name, index=False, encoding="utf-8-sig")
    hard_enriched.to_csv(HARD_DIR / hard_csv.name, index=False, encoding="utf-8-sig")
    write_xlsx(rep_enriched, rep_xlsx, "representative_final")
    write_xlsx(hard_enriched, hard_xlsx, "hardcase_final")
    write_xlsx(rep_enriched, REP_DIR / rep_xlsx.name, "representative_final")
    write_xlsx(hard_enriched, HARD_DIR / hard_xlsx.name, "hardcase_final")

    guidelines = guidelines_text()
    (ROOT / "annotation_guidelines_final.md").write_text(guidelines, encoding="utf-8")
    (PKG / "annotation_guidelines_final.md").write_text(guidelines, encoding="utf-8")

    readme = build_readme(rep_source_missing, rep_target_missing, hard_source_missing, hard_target_missing)
    (ROOT / "README_annotation_final.txt").write_text(readme, encoding="utf-8")
    (PKG / "README_annotation_final.txt").write_text(readme, encoding="utf-8")

    print(
        f"Prepared final annotation package: representative={len(rep_enriched)}, hardcase={len(hard_enriched)}, "
        f"rep_source_missing={len(rep_source_missing)}, hard_source_missing={len(hard_source_missing)}"
    )


if __name__ == "__main__":
    main()
