#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import importlib.util
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
RESULTS = ROOT / "results"
RESULTS.mkdir(parents=True, exist_ok=True)


def load_t611():
    path = ROOT / "task6.11_rebuild_target_aligned_contexts.py"
    spec = importlib.util.spec_from_file_location("t611", path)
    module = importlib.util.module_from_spec(spec)
    sys.modules["t611"] = module
    spec.loader.exec_module(module)
    return module


t611 = load_t611()


CURRENT_PATH = ROOT / "target_aligned_contexts.csv"
V22_PATH = ROOT / "target_aligned_contexts_v22_candidate.csv"
AMBIGUOUS_XLSX = ROOT / "ambiguous_edges_for_manual_alignment_v21.xlsx"


FAILURE_REASON_ORDER = [
    "missing_pdf_or_text",
    "reference_section_not_found",
    "target_reference_entry_not_matched",
    "citation_marker_not_found",
    "author_year_format_not_supported",
    "title_fuzzy_match_failed",
    "grouped_or_range_uncertain",
    "poor_pdf_text_extraction",
    "unknown",
]

STOPWORDS = {
    "a", "an", "and", "the", "of", "for", "on", "in", "to", "from", "with", "by",
    "based", "using", "under", "towards", "toward", "study", "survey", "analysis",
}


@dataclass
class SourceMeta:
    pdf_exists: bool
    pdf_name: str
    full_text_len: int
    refs_text_len: int
    has_refs_section: bool
    numbered_entries: int
    author_year_entries: int
    poor_text_flag: bool


def normalize_loose(text: object) -> str:
    if pd.isna(text):
        return ""
    s = unicodedata.normalize("NFKD", str(text)).lower()
    s = s.replace("ﬁ", "fi").replace("ﬂ", "fl").replace("", "-").replace("—", "-").replace("–", "-")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def normalize_compact(text: object) -> str:
    return normalize_loose(text).replace(" ", "")


def core_title(text: object) -> str:
    raw = str(text or "")
    head = re.split(r"[:\-–—]", raw, maxsplit=1)[0]
    return normalize_loose(head)


def token_set(text: object) -> set[str]:
    return {tok for tok in normalize_loose(text).split() if tok and tok not in STOPWORDS and len(tok) > 2}


def title_score_v22(target_title: str, entry_text: str) -> Tuple[float, str]:
    target_norm = normalize_loose(target_title)
    entry_norm = normalize_loose(entry_text)
    target_compact = normalize_compact(target_title)
    entry_compact = normalize_compact(entry_text)
    target_core = core_title(target_title)
    entry_core = core_title(entry_text)
    tokens_a = token_set(target_title)
    tokens_b = token_set(entry_text)
    token_recall = len(tokens_a & tokens_b) / max(1, len(tokens_a))
    compact_ratio = t611.SequenceMatcher(None, target_compact, entry_compact).ratio() if target_compact and entry_compact else 0.0
    core_ratio = t611.SequenceMatcher(None, target_core.replace(" ", ""), entry_core.replace(" ", "")).ratio() if target_core and entry_core else 0.0
    loose_ratio = t611.SequenceMatcher(None, target_norm, entry_norm).ratio() if target_norm and entry_norm else 0.0

    if target_compact and target_compact in entry_compact:
        return 0.99, "title_compact_exact"
    if target_core and target_core.replace(" ", "") and target_core.replace(" ", "") in entry_compact:
        return 0.96, "core_title_exact"
    if compact_ratio >= 0.92 and token_recall >= 0.60:
        return max(compact_ratio, 0.92), "title_compact_fuzzy_high"
    if core_ratio >= 0.90 and token_recall >= 0.55:
        return max(core_ratio, 0.90), "core_title_fuzzy_high"
    score = max(0.55 * loose_ratio + 0.25 * compact_ratio + 0.20 * token_recall, core_ratio * 0.95)
    return float(score), "title_fuzzy_candidate"


def parse_author_year_entries(refs_text: str) -> List[t611.ReferenceEntry]:
    if not refs_text:
        return []
    cleaned = refs_text.replace("\r", "\n")
    pattern = re.compile(r"(?ms)(?:^|\n)([A-Z][^\n]{0,200}?(?:19|20)\d{2}\..*?)(?=(?:\n[A-Z][^\n]{0,200}?(?:19|20)\d{2}\.)|\Z)")
    entries: List[t611.ReferenceEntry] = []
    idx = 1
    for match in pattern.finditer(cleaned):
        entry = re.sub(r"\s+", " ", match.group(1)).strip()
        if len(entry) >= 40:
            entries.append(t611.ReferenceEntry(f"AY{idx}", None, entry))
            idx += 1
    return entries


def extract_first_author_surname(entry_text: str, fallback: str = "") -> str:
    text = str(entry_text or "")
    if fallback:
        fb = normalize_loose(fallback).split()
        if fb:
            return fb[-1]
    m = re.match(r"\s*([A-Z][a-zA-Z'`-]+)", text)
    if m:
        return normalize_loose(m.group(1))
    return ""


def extract_author_year_mentions(body_text: str, author_name: str, year: Optional[int]) -> List[str]:
    if not body_text or not author_name or not year:
        return []
    surname = extract_first_author_surname("", author_name)
    if not surname:
        return []
    spans = t611.sentence_boundaries(body_text)
    mentions: List[str] = []
    patts = [
        re.compile(rf"\b{re.escape(surname)}\b[^.()\\n]{{0,40}}(?:19|20)\d{{2}}", re.I),
        re.compile(rf"\b{re.escape(surname)}\b\s*(?:et al\.?|and [A-Z][a-zA-Z'`-]+)?\s*\((?:19|20)\d{{2}}\)", re.I),
        re.compile(rf"\((?:[^)]*?)\b{re.escape(surname)}\b(?:[^)]*?)(?:19|20)\d{{2}}(?:[^)]*?)\)", re.I),
    ]
    target_year = str(year)
    for idx, (start, end) in enumerate(spans):
        sentence = body_text[start:end]
        if target_year not in sentence:
            continue
        if not any(p.search(sentence) for p in patts):
            continue
        left = spans[idx - 1][0] if idx > 0 else start
        right = spans[idx + 1][1] if idx + 1 < len(spans) else end
        ctx = re.sub(r"\s+", " ", body_text[left:right]).strip()
        if ctx and ctx not in mentions:
            mentions.append(ctx)
    return mentions


def infer_source_meta(papers: pd.DataFrame) -> Tuple[Dict[str, Path], Dict[str, SourceMeta]]:
    pdf_map = t611.build_pdf_path_map(papers)
    meta: Dict[str, SourceMeta] = {}
    for paper_id in papers["paper_id"]:
        pdf = pdf_map.get(paper_id)
        if pdf is None:
            meta[paper_id] = SourceMeta(False, "", 0, 0, False, 0, 0, False)
            continue
        text = t611.extract_pdf_text(pdf)
        body, refs = t611.split_body_and_references(text)
        numbered = t611.parse_reference_entries(refs)
        ay_entries = parse_author_year_entries(refs) if not numbered else []
        odd_ratio = sum(ch == "�" for ch in text) / max(1, len(text))
        poor_flag = (len(text) < 2000) or odd_ratio > 0.01
        meta[paper_id] = SourceMeta(
            True,
            pdf.name,
            len(text),
            len(refs),
            bool(refs),
            len(numbered),
            len(ay_entries),
            poor_flag,
        )
    return pdf_map, meta


def classify_failure_reason(row: pd.Series, meta: SourceMeta) -> str:
    status = str(row.get("alignment_status", ""))
    method = str(row.get("reference_match_method", ""))
    num_mentions = int(pd.to_numeric(row.get("num_mentions", 0), errors="coerce") or 0)
    target_entry = str(row.get("target_reference_entry", "") or "").strip()
    marker = str(row.get("target_reference_marker", "") or "").strip()
    ctype = str(row.get("context_type", "") or "").strip().lower()

    if not meta.pdf_exists:
        return "missing_pdf_or_text"
    if meta.poor_text_flag and (meta.full_text_len < 2000 or not meta.has_refs_section):
        return "poor_pdf_text_extraction"
    if not meta.has_refs_section:
        return "reference_section_not_found"
    if method == "no_reference_entries":
        if meta.author_year_entries > 0:
            return "author_year_format_not_supported"
        if meta.numbered_entries == 0:
            return "reference_section_not_found"
    if status == "ambiguous" and target_entry and num_mentions == 0:
        return "citation_marker_not_found"
    if status == "ambiguous" and ctype in {"grouped", "range", "mixed"}:
        return "grouped_or_range_uncertain"
    if method == "failed":
        score = float(pd.to_numeric(row.get("reference_match_score", 0.0), errors="coerce") or 0.0)
        if target_entry and marker and num_mentions > 0:
            return "title_fuzzy_match_failed"
        if target_entry:
            return "citation_marker_not_found" if score >= 0.75 else "title_fuzzy_match_failed"
        return "target_reference_entry_not_matched"
    if method == "pdf_missing":
        return "missing_pdf_or_text"
    return "unknown"


def build_failure_tables(aligned: pd.DataFrame, meta_map: Dict[str, SourceMeta]) -> pd.DataFrame:
    sub = aligned[aligned["alignment_status"].isin(["ambiguous", "failed"])].copy()
    sub["failure_reason"] = sub.apply(lambda r: classify_failure_reason(r, meta_map.get(r["source_id"], SourceMeta(False, "", 0, 0, False, 0, 0, False))), axis=1)
    sub["pdf_name"] = sub["source_id"].map(lambda sid: meta_map.get(sid).pdf_name if sid in meta_map else "")
    reason_table = (
        sub.groupby("failure_reason", as_index=False)
        .agg(edge_count=("edge_id", "size"), source_paper_count=("source_id", "nunique"))
        .sort_values(["edge_count", "failure_reason"], ascending=[False, True])
    )
    reason_table = (
        pd.DataFrame({"failure_reason": FAILURE_REASON_ORDER})
        .merge(reason_table, on="failure_reason", how="left")
        .fillna({"edge_count": 0, "source_paper_count": 0})
    )
    reason_table["edge_count"] = reason_table["edge_count"].astype(int)
    reason_table["source_paper_count"] = reason_table["source_paper_count"].astype(int)
    reason_table.to_csv(RESULTS / "table_alignment_failed_reason_v21.csv", index=False, encoding="utf-8-sig")

    total_fail = len(sub)
    lines = [
        "# Alignment Failed Reason Summary v2.1",
        "",
        f"- Total ambiguous/failed edges: `{total_fail}`",
        "",
        md(reason_table),
        "",
        "## Interpretation",
        "",
        "- `missing_pdf_or_text` means the source PDF is not locally available, so coverage cannot be improved without adding source files.",
        "- `reference_section_not_found` and `poor_pdf_text_extraction` indicate upstream text availability or PDF parsing limits.",
        "- `author_year_format_not_supported` identifies sources where references and body citations likely use author-year rather than numbered markers.",
        "- `citation_marker_not_found` means the target reference entry was matched, but the body marker could not be aligned confidently.",
        "- `title_fuzzy_match_failed` marks cases where a candidate reference marker exists but title matching remains too weak for formal acceptance.",
    ]
    (RESULTS / "alignment_failed_reason_summary_v21.md").write_text("\n".join(lines), encoding="utf-8")

    source_stats = (
        sub.groupby(["source_id", "citing_paper_title"], as_index=False)
        .agg(
            failed_edges=("alignment_status", lambda s: int((s == "failed").sum())),
            ambiguous_edges=("alignment_status", lambda s: int((s == "ambiguous").sum())),
            main_failure_reason=("failure_reason", lambda s: s.value_counts().idxmax()),
        )
    )
    total_from_source = aligned.groupby(["source_id", "citing_paper_title"], as_index=False).size().rename(columns={"size": "total_edges_from_source"})
    source_stats = total_from_source.merge(source_stats, on=["source_id", "citing_paper_title"], how="left").fillna({"failed_edges": 0, "ambiguous_edges": 0, "main_failure_reason": ""})
    source_stats["failed_edges"] = source_stats["failed_edges"].astype(int)
    source_stats["ambiguous_edges"] = source_stats["ambiguous_edges"].astype(int)
    source_stats["failure_ratio"] = (source_stats["failed_edges"] + source_stats["ambiguous_edges"]) / source_stats["total_edges_from_source"].clip(lower=1)
    source_stats = source_stats.sort_values(["failed_edges", "ambiguous_edges", "failure_ratio"], ascending=[False, False, False])
    source_stats.to_csv(RESULTS / "table_failed_edges_by_source_paper_v21.csv", index=False, encoding="utf-8-sig")
    return sub


def write_ambiguous_review_xlsx(aligned: pd.DataFrame) -> None:
    amb = aligned[aligned["alignment_status"] == "ambiguous"].copy()
    review = pd.DataFrame(
        {
            "edge_id": amb["edge_id"],
            "citing_paper_title": amb["citing_paper_title"],
            "cited_paper_title": amb["cited_paper_title"],
            "target_reference_entry": amb["target_reference_entry"],
            "candidate_reference_marker": amb["target_reference_marker"],
            "extracted_context": amb["all_target_aligned_contexts"],
            "auto_alignment_status": amb["alignment_status"],
            "auto_alignment_reason": amb["alignment_note"],
            "human_alignment_status": "",
            "human_reference_marker": "",
            "human_note": "",
        }
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "ambiguous_review"
    ws.append(list(review.columns))
    control_re = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
    for _, row in review.iterrows():
        ws.append([control_re.sub("", "" if pd.isna(v) else str(v)) for v in row.values])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    input_cols = {"human_alignment_status", "human_reference_marker", "human_note"}
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, name in enumerate(review.columns, start=1):
        width = 18
        if name in {"citing_paper_title", "cited_paper_title", "target_reference_entry"}:
            width = 34
        if name in {"extracted_context", "human_note", "auto_alignment_reason"}:
            width = 56
        ws.column_dimensions[get_column_letter(idx)].width = width
        if name in input_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, idx).fill = input_fill
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 48
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(AMBIGUOUS_XLSX)


def candidate_rescue(aligned: pd.DataFrame, papers: pd.DataFrame, meta_map: Dict[str, SourceMeta]) -> pd.DataFrame:
    authorships = t611.safe_read_csv(ROOT / "authorships_network.csv") if (ROOT / "authorships_network.csv").exists() else pd.DataFrame()
    first_author_map = t611.load_first_author_map(authorships)
    title_map = dict(zip(papers["paper_id"], papers["title"]))
    year_map = dict(zip(papers["paper_id"], papers["year"]))
    pdf_map = t611.build_pdf_path_map(papers)

    out = aligned.copy()
    out["candidate_upgrade_reason"] = ""

    for idx, row in out.iterrows():
        if row["alignment_status"] in {"exact", "high_confidence", "grouped", "range"}:
            continue
        source_id = row["source_id"]
        target_id = row["target_id"]
        pdf = pdf_map.get(source_id)
        if pdf is None:
            continue

        full_text = t611.extract_pdf_text(pdf)
        body_text, refs_text = t611.split_body_and_references(full_text)
        numbered_entries = t611.parse_reference_entries(refs_text)
        ay_entries = parse_author_year_entries(refs_text) if not numbered_entries else []
        entries = numbered_entries if numbered_entries else ay_entries
        if not entries:
            continue

        best_entry = None
        best_method = ""
        best_score = 0.0
        for entry in entries:
            score, method = title_score_v22(title_map.get(target_id, ""), entry.entry_text)
            author = first_author_map.get(target_id, "")
            author_hit = extract_first_author_surname(entry.entry_text, author) in normalize_loose(entry.entry_text)
            year_hit = str(int(year_map.get(target_id, 0))) in entry.entry_text if year_map.get(target_id) else False
            if author_hit and year_hit:
                score = max(score, min(0.91, score + 0.08))
            if score > best_score:
                best_entry, best_method, best_score = entry, method, score

        if best_entry is None:
            continue

        mentions: List[str] = []
        context_types: List[str] = []
        if best_entry.marker_number is not None:
            mentions, context_types = t611.extract_target_mentions(body_text, best_entry.marker_number)
        else:
            mentions = extract_author_year_mentions(body_text, first_author_map.get(target_id, ""), t611.parse_year(year_map.get(target_id)))
            context_types = ["author_year"] * len(mentions)

        if not mentions:
            continue

        strong = best_score >= 0.90
        moderate = best_score >= 0.84 and ("author_year" in context_types or row["reference_match_score"] >= 0.50)
        if not (strong or moderate):
            continue

        new_status = "high_confidence"
        note_reason = f"Candidate v2.2 rescue via {best_method} (score={best_score:.3f})"
        if any(ct == "range" for ct in context_types):
            new_status = "range"
        elif any(ct == "grouped" for ct in context_types):
            new_status = "grouped"

        out.at[idx, "target_reference_marker"] = best_entry.marker
        out.at[idx, "target_reference_entry"] = best_entry.entry_text
        out.at[idx, "reference_match_method"] = best_method
        out.at[idx, "reference_match_score"] = float(best_score)
        out.at[idx, "num_mentions"] = len(mentions)
        out.at[idx, "all_target_aligned_contexts"] = "\n\n... [another target-aligned mention] ...\n\n".join(mentions)
        out.at[idx, "context_type"] = "mixed" if len(set(context_types)) > 1 else context_types[0]
        out.at[idx, "alignment_status"] = new_status
        out.at[idx, "alignment_note"] = note_reason
        out.at[idx, "candidate_upgrade_reason"] = note_reason

    out.to_csv(V22_PATH, index=False, encoding="utf-8-sig")
    return out


def compare_coverage(v21: pd.DataFrame, v22: pd.DataFrame) -> None:
    def row(df: pd.DataFrame, version: str) -> Dict[str, object]:
        statuses = df["alignment_status"].value_counts()
        scored = int(df["alignment_status"].isin(["exact", "high_confidence", "grouped", "range"]).sum())
        return {
            "version": version,
            "high_confidence": int(statuses.get("high_confidence", 0)),
            "grouped": int(statuses.get("grouped", 0)),
            "range": int(statuses.get("range", 0)),
            "ambiguous": int(statuses.get("ambiguous", 0)),
            "failed": int(statuses.get("failed", 0)),
            "scored_edges": scored,
            "coverage_ratio": scored / max(1, len(df)),
        }

    comp = pd.DataFrame([row(v21, "v2.1"), row(v22, "v2.2_candidate")])
    comp.to_csv(RESULTS / "table_alignment_coverage_compare_v21_v22.csv", index=False, encoding="utf-8-sig")


def md(df: pd.DataFrame) -> str:
    show = df.fillna("")
    lines = [
        "| " + " | ".join(map(str, show.columns)) + " |",
        "| " + " | ".join(["---"] * len(show.columns)) + " |",
    ]
    for _, row in show.iterrows():
        vals = []
        for val in row:
            if isinstance(val, float):
                vals.append(f"{val:.6f}")
            else:
                vals.append(str(val).replace("\n", " ").strip())
        lines.append("| " + " | ".join(vals) + " |")
    return "\n".join(lines)


def write_candidate_summary(v21: pd.DataFrame, v22: pd.DataFrame, failed_sub: pd.DataFrame) -> None:
    comp = pd.read_csv(RESULTS / "table_alignment_coverage_compare_v21_v22.csv")
    upgraded = v22.merge(v21[["edge_id", "alignment_status"]].rename(columns={"alignment_status": "old_status"}), on="edge_id", how="left")
    upgraded = upgraded[(upgraded["old_status"].isin(["ambiguous", "failed"])) & (upgraded["alignment_status"].isin(["high_confidence", "grouped", "range", "exact"]))]
    reason_counts = failed_sub["failure_reason"].value_counts().to_dict()
    lines = [
        "# v2.2 Candidate Coverage Summary",
        "",
        md(comp),
        "",
        f"- Candidate rescues: `{len(upgraded)}` edges.",
        f"- Largest existing bottlenecks remain `missing_pdf_or_text` (`{reason_counts.get('missing_pdf_or_text', 0)}` edges) and `reference_section_not_found` / `poor_pdf_text_extraction` (`{reason_counts.get('reference_section_not_found', 0) + reason_counts.get('poor_pdf_text_extraction', 0)}` edges combined).",
        "- The candidate file is not promoted to the formal semantic layer automatically. It should be used only after manual review of rescued edges, especially those upgraded from fuzzy title matching or author-year support.",
    ]
    (RESULTS / "alignment_failed_reason_summary_v21.md").write_text(
        (RESULTS / "alignment_failed_reason_summary_v21.md").read_text(encoding="utf-8") + "\n\n## v2.2 candidate note\n\n" + "\n".join(lines),
        encoding="utf-8",
    )


def main() -> None:
    aligned = pd.read_csv(CURRENT_PATH)
    papers = t611.prepare_papers(t611.safe_read_csv(ROOT / "all_connected_papers.csv"))
    _, meta_map = infer_source_meta(papers)
    failed_sub = build_failure_tables(aligned, meta_map)
    write_ambiguous_review_xlsx(aligned)
    v22 = candidate_rescue(aligned, papers, meta_map)
    compare_coverage(aligned, v22)
    write_candidate_summary(aligned, v22, failed_sub)
    print("Wrote:", RESULTS / "table_alignment_failed_reason_v21.csv")
    print("Wrote:", RESULTS / "table_failed_edges_by_source_paper_v21.csv")
    print("Wrote:", AMBIGUOUS_XLSX)
    print("Wrote:", V22_PATH)
    print("Wrote:", RESULTS / "table_alignment_coverage_compare_v21_v22.csv")


if __name__ == "__main__":
    main()
