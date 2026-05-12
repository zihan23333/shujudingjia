#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
PKG = ROOT / "llm_reliability_annotation_package_v2"
INPUT_CONTEXTS = ROOT / "target_aligned_contexts.csv"
INPUT_LLM = ROOT / "llm_results_target_aligned_v2.csv"
CUE_AUDIT = ROOT / "results" / "table_sentiment_cue_audit.csv"


def normalize_section(value: object) -> str:
    text = str(value).strip().lower()
    if "method" in text:
        return "Methodology"
    if "result" in text:
        return "Result"
    if "discussion" in text:
        return "Discussion"
    if "conclusion" in text:
        return "Conclusion"
    if "introduction" in text or "related" in text or "background" in text:
        return "Introduction"
    return "Other"


def stratified_sample(df: pd.DataFrame, n: int = 50) -> pd.DataFrame:
    work = df.copy()
    work["section_bucket"] = work["LLM_section"].apply(normalize_section)
    work["relevance_bucket"] = pd.cut(work["LLM_relevance"], bins=[-0.01, 0.4, 0.7, 1.01], labels=["low", "mid", "high"])
    work["q_bucket"] = pd.cut(work["LLM_q"], bins=[-0.01, 0.08, 0.22, 1.01], labels=["low", "mid", "high"])
    work["stratum"] = (
        work["alignment_status"].astype(str)
        + "|"
        + work["section_bucket"].astype(str)
        + "|"
        + work["relevance_bucket"].astype(str)
        + "|"
        + work["q_bucket"].astype(str)
    )

    picks = []
    manual = work[work["manual_check_needed"] == 1].drop_duplicates("edge_id")
    if not manual.empty:
        picks.append(manual)
    for _, group in work.groupby("stratum", sort=False):
        picks.append(group.sample(n=1, random_state=42))

    sampled = pd.concat(picks, ignore_index=True).drop_duplicates("edge_id")
    if len(sampled) < min(n, len(work)):
        remain = work[~work["edge_id"].isin(sampled["edge_id"])].copy()
        extra = remain.sample(n=min(n - len(sampled), len(remain)), random_state=42)
        sampled = pd.concat([sampled, extra], ignore_index=True)

    sampled = sampled.head(min(n, len(work))).reset_index(drop=True)
    sampled.insert(0, "sample_id", [f"S{i:03d}" for i in range(1, len(sampled) + 1)])
    return sampled


def write_xlsx(df: pd.DataFrame, path: Path) -> None:
    control_re = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
    wb = Workbook()
    ws = wb.active
    ws.title = "annotation_v2"
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([control_re.sub("", "" if pd.isna(v) else str(v)) for v in row.values])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    input_cols = {
        "human_alignment_check",
        "human_primary_section",
        "human_sentiment",
        "human_relevance",
        "annotation_note",
    }

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, name in enumerate(df.columns, start=1):
        width = 18
        if name in {"citing_paper_title", "cited_paper_title", "target_reference_entry"}:
            width = 34
        if name in {"all_target_aligned_contexts", "annotation_note"}:
            width = 60
        ws.column_dimensions[get_column_letter(idx)].width = width
        if name in input_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, idx).fill = input_fill

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 52
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(path)


def build_guidelines() -> str:
    return """# 人工标注说明 v2

本次标注只针对新的 DeepSeek target-aligned semantic layer。每一行是一条论文到论文的 citation edge。请只评价当前 `cited_paper_title` 对应的 target cited paper。

## 标注原则

1. 阅读 `all_target_aligned_contexts` 中提供的全部目标对齐上下文。
2. 只评价当前 target cited paper 的语义作用，不评价 grouped / range 引用中的其他文献。
3. 如果是 grouped / range citation，需要根据当前 target 在该上下文中的作用判断。
4. 如果人工认为当前 context 仍未真正对准 target，请在 `human_alignment_check` 中填写 `ambiguous` 或 `wrong`。

## 需要填写的字段

- `human_alignment_check`
- `human_primary_section`
- `human_sentiment`
- `human_relevance`
- `annotation_note`

## 标签说明

- `human_alignment_check`：`correct / ambiguous / wrong`
- `human_primary_section`：`Introduction / Methodology / Result / Discussion / Conclusion / Other`
- `human_sentiment`：`positive / neutral / negative`
- `human_relevance`：`0 / 0.25 / 0.5 / 0.75 / 1.0`
- `annotation_note`：记录 mixed usage、target 对齐疑问或其他备注

Reliability analysis v2 只在 `human_alignment_check = correct` 的样本上计算。
"""


def build_readme() -> str:
    return """LLM reliability annotation package v2

优先使用：
- sampled_contexts_for_human_annotation_v2.xlsx
- annotation_guidelines_v2.md

说明：
- 本包只针对 DeepSeek target-aligned semantic layer v2.1
- 标注者只评价当前 target cited paper
- 如果认为上下文仍未对准 target，请填写 human_alignment_check = ambiguous 或 wrong
- reliability 只在 human_alignment_check = correct 的样本上计算
- manual_check_needed = 1 的样本建议优先人工复核
"""


def main() -> None:
    PKG.mkdir(parents=True, exist_ok=True)
    contexts = pd.read_csv(INPUT_CONTEXTS)
    llm = pd.read_csv(INPUT_LLM)

    merged = contexts.merge(
        llm[
            [
                "edge_id",
                "section",
                "sentiment",
                "relevance",
                "confidence",
                "q_ij",
            ]
        ].rename(
            columns={
                "section": "LLM_section",
                "sentiment": "LLM_sentiment",
                "relevance": "LLM_relevance",
                "confidence": "LLM_confidence",
                "q_ij": "LLM_q",
            }
        ),
        on="edge_id",
        how="inner",
    )
    merged = merged[merged["alignment_status"].isin(["high_confidence", "grouped", "range"])].copy()
    if CUE_AUDIT.exists():
        cue = pd.read_csv(CUE_AUDIT)[["edge_id", "manual_check_needed"]].drop_duplicates("edge_id")
        merged = merged.merge(cue, on="edge_id", how="left")
        merged["manual_check_needed"] = pd.to_numeric(merged["manual_check_needed"], errors="coerce").fillna(0).astype(int)
    else:
        merged["manual_check_needed"] = 0

    sampled = stratified_sample(
        merged[
            [
                "edge_id",
                "citing_paper_title",
                "cited_paper_title",
                "alignment_status",
                "target_reference_marker",
                "target_reference_entry",
                "all_target_aligned_contexts",
                "LLM_section",
                "LLM_sentiment",
                "LLM_relevance",
                "LLM_confidence",
                "LLM_q",
                "manual_check_needed",
            ]
        ],
        n=50,
    )
    sampled["human_alignment_check"] = ""
    sampled["human_primary_section"] = ""
    sampled["human_sentiment"] = ""
    sampled["human_relevance"] = ""
    sampled["annotation_note"] = ""
    sampled = sampled[
        [
            "sample_id",
            "edge_id",
            "citing_paper_title",
            "cited_paper_title",
            "alignment_status",
            "target_reference_marker",
            "target_reference_entry",
            "all_target_aligned_contexts",
            "LLM_section",
            "LLM_sentiment",
            "LLM_relevance",
            "LLM_confidence",
            "LLM_q",
            "manual_check_needed",
            "human_alignment_check",
            "human_primary_section",
            "human_sentiment",
            "human_relevance",
            "annotation_note",
        ]
    ]

    csv_targets = [
        ROOT / "sampled_contexts_for_human_annotation_v2.csv",
        PKG / "sampled_contexts_for_human_annotation_v2.csv",
    ]
    xlsx_targets = [
        ROOT / "sampled_contexts_for_human_annotation_v2.xlsx",
        PKG / "sampled_contexts_for_human_annotation_v2.xlsx",
    ]
    for path in csv_targets:
        sampled.to_csv(path, index=False, encoding="utf-8-sig")
    for path in xlsx_targets:
        write_xlsx(sampled, path)

    for path in [ROOT / "annotation_guidelines_v2.md", PKG / "annotation_guidelines_v2.md"]:
        path.write_text(build_guidelines(), encoding="utf-8")
    for path in [ROOT / "README_annotation_v2.txt", PKG / "README_annotation_v2.txt"]:
        path.write_text(build_readme(), encoding="utf-8")

    print(f"Prepared v2 annotation package with {len(sampled)} samples.")


if __name__ == "__main__":
    main()
