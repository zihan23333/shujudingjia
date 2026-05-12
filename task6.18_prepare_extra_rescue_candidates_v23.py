#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import importlib.util
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
RESULTS = ROOT / "results"
RESULTS.mkdir(parents=True, exist_ok=True)


def load_audit_module():
    path = ROOT / "task6.11_alignment_failure_audit_and_v22_candidate.py"
    spec = importlib.util.spec_from_file_location("auditv23", path)
    module = importlib.util.module_from_spec(spec)
    sys.modules["auditv23"] = module
    spec.loader.exec_module(module)
    return module


audit = load_audit_module()


V22_PATH = ROOT / "target_aligned_contexts_v22.csv"
OUT_XLSX = ROOT / "extra_rescue_candidates_for_manual_review_v23.xlsx"
OUT_CSV = ROOT / "extra_rescue_candidates_for_manual_review_v23.csv"
SUMMARY_MD = ROOT / "extra_rescue_v23_summary.md"


def write_xlsx(df: pd.DataFrame, path: Path) -> None:
    control_re = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
    wb = Workbook()
    ws = wb.active
    ws.title = "extra_rescue_v23"
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([control_re.sub("", "" if pd.isna(v) else str(v)) for v in row.values])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    input_cols = {"human_alignment_status", "human_reference_marker_confirmed", "human_note"}
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, name in enumerate(df.columns, start=1):
        width = 18
        if name in {"citing_paper_title", "cited_paper_title", "target_reference_entry"}:
            width = 36
        if name in {"all_target_aligned_contexts", "candidate_reason", "human_note"}:
            width = 60
        ws.column_dimensions[get_column_letter(idx)].width = width
        if name in input_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, idx).fill = input_fill
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 52
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    v22 = pd.read_csv(V22_PATH)
    papers = audit.t611.prepare_papers(audit.t611.safe_read_csv(ROOT / "all_connected_papers.csv"))
    _, meta_map = audit.infer_source_meta(papers)
    sub = v22[v22["alignment_status"].isin(["failed", "ambiguous"])].copy()
    sub["failure_reason_v21"] = sub.apply(
        lambda r: audit.classify_failure_reason(
            r,
            meta_map.get(r["source_id"], audit.SourceMeta(False, "", 0, 0, False, 0, 0, False)),
        ),
        axis=1,
    )
    eligible = sub[sub["failure_reason_v21"].isin(["title_fuzzy_match_failed", "citation_marker_not_found"])].copy()
    eligible[["title_score_check", "title_score_method"]] = eligible.apply(
        lambda r: pd.Series(audit.title_score_v22(r["cited_paper_title"], r["target_reference_entry"])),
        axis=1,
    )

    # Very conservative one-round extra rescue:
    # keep only rows with
    # - title_fuzzy_match_failed
    # - at least one extracted mention
    # - reasonable title consistency under OCR noise
    # - explicit numeric marker form
    candidates = eligible[
        (eligible["failure_reason_v21"] == "title_fuzzy_match_failed")
        & (pd.to_numeric(eligible["num_mentions"], errors="coerce").fillna(0) >= 1)
        & (pd.to_numeric(eligible["title_score_check"], errors="coerce").fillna(0) >= 0.70)
        & eligible["target_reference_marker"].astype(str).str.contains(r"\d", regex=True)
        & eligible["all_target_aligned_contexts"].fillna("").astype(str).str.len().ge(40)
    ].copy()

    candidates["candidate_alignment_status"] = "high_confidence"
    candidates["candidate_reason"] = candidates.apply(
        lambda r: (
            f"Single-round extra rescue candidate: v2.1 failure type `{r['failure_reason_v21']}`; "
            f"body still contains an extracted target-aligned mention for marker `{r['target_reference_marker']}`; "
            f"title consistency under OCR-normalized matching is `{float(r['title_score_check']):.3f}` via `{r['title_score_method']}`. "
            "This candidate should be accepted only if manual review confirms that the reference entry and the extracted context both truly point to the current target paper."
        ),
        axis=1,
    )

    out = candidates[
        [
            "edge_id",
            "source_id",
            "target_id",
            "citing_paper_title",
            "cited_paper_title",
            "failure_reason_v21",
            "candidate_alignment_status",
            "target_reference_marker",
            "target_reference_entry",
            "title_score_method",
            "title_score_check",
            "all_target_aligned_contexts",
            "candidate_reason",
        ]
    ].rename(
        columns={
            "title_score_method": "reference_match_method",
            "title_score_check": "reference_match_score",
        }
    ).copy()
    out["human_alignment_status"] = ""
    out["human_reference_marker_confirmed"] = ""
    out["human_note"] = ""
    out.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
    write_xlsx(out, OUT_XLSX)

    candidate_count = len(out)
    correct_count = 0
    ambiguous_count = 0
    wrong_count = 0
    final_scored = int(v22["alignment_status"].isin(["exact", "high_confidence", "grouped", "range"]).sum())
    final_ratio = final_scored / max(1, len(v22))
    recommend = "No"
    if candidate_count >= 8:
        recommend = "Pending manual review"

    lines = [
        "# Extra Rescue v2.3 Summary",
        "",
        f"- Candidate count: `{candidate_count}`",
        f"- Human confirmed `correct`: `{correct_count}`",
        f"- Human confirmed `ambiguous`: `{ambiguous_count}`",
        f"- Human confirmed `wrong`: `{wrong_count}`",
        f"- Current formal scored edges remain: `{final_scored}`",
        f"- Current formal coverage ratio remains: `{final_ratio:.2%}`",
        f"- Suggest upgrade to v2.3 now: `{recommend}`",
        "",
        "Interpretation:",
        "- This is intentionally a single, conservative extra-rescue round on top of the formal v2.2 layer.",
        "- Only edges from `title_fuzzy_match_failed` and `citation_marker_not_found` were considered.",
        "- `missing_pdf_or_text` was not revisited.",
        "- If manual review later confirms at least 8 additional edges as `correct`, then a formal `target_aligned_contexts_v23.csv` upgrade can be justified; otherwise the current 112-edge v2.2 layer should remain the formal semantic layer.",
    ]
    SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"Prepared {candidate_count} extra rescue candidates for manual review.")


if __name__ == "__main__":
    main()
